"""
xlsx_generator.py — openpyxl 기반 XLSX 생성기
FPOF Document Converter
"""
from __future__ import annotations
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..core.content_model import (
    DocumentContent, Section, Block, BlockType, BrandTheme, TableRow
)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color.lstrip('#'))


def _font(size: int = 11, bold: bool = False, color: str = "#000000") -> Font:
    return Font(size=size, bold=bold, color=color.lstrip('#'))


def _thin_border() -> Border:
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)


def _write_cover_sheet(wb: Workbook, doc: DocumentContent, brand: BrandTheme):
    ws = wb.active
    ws.title = "Cover"

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50

    ws['A1'] = brand.brand_name.upper()
    ws['A1'].font = Font(size=14, bold=True, color=brand.primary.lstrip('#'))

    ws['A2'] = doc.title
    ws['A2'].font = Font(size=18, bold=True)

    if doc.subtitle:
        ws['A3'] = doc.subtitle
        ws['A3'].font = Font(size=12, color=brand.text_light.lstrip('#'))

    meta_parts = [x for x in [doc.season, str(doc.date)] if x]
    if meta_parts:
        ws['A4'] = "  ·  ".join(meta_parts)
        ws['A4'].font = Font(size=10, color='888888')

    # 메타데이터 테이블
    row = 6
    meta_items = [
        ("시즌", doc.season),
        ("날짜", str(doc.date)),
        ("작성자", doc.author),
        ("문서 유형", doc.doc_type),
    ]
    for label, value in meta_items:
        if value:
            ws.cell(row, 1, label).font = _font(10, True, brand.primary)
            ws.cell(row, 2, value).font = _font(10)
            row += 1


def _write_section_sheet(wb: Workbook, section: Section, brand: BrandTheme,
                          sheet_idx: int):
    # 시트명 정리 (Excel 31자 제한, 특수문자 제거)
    import re
    raw_name = re.sub(r'[\\/*?:\[\]]', '', section.title)[:30] or f"Sheet{sheet_idx}"
    ws = wb.create_sheet(title=raw_name)

    row = 1

    # 시트 제목
    ws.cell(row, 1, section.title).font = Font(size=14, bold=True, color=brand.secondary.lstrip('#'))
    ws.cell(row, 1).fill = _fill(brand.primary)
    ws.cell(row, 1).font = Font(size=14, bold=True, color=brand.background.lstrip('#'))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 28
    row += 2

    for block in section.blocks:
        if block.type == BlockType.HEADING3:
            ws.cell(row, 1, block.text).font = Font(size=12, bold=True, color=brand.primary.lstrip('#'))
            row += 1

        elif block.type == BlockType.PARAGRAPH:
            if block.text.strip():
                ws.cell(row, 1, block.text).alignment = Alignment(wrap_text=True)
                ws.cell(row, 1).font = _font(10)
                ws.row_dimensions[row].height = 20
                row += 1

        elif block.type == BlockType.BULLET:
            for item in block.items:
                ws.cell(row, 1, f"• {item}").font = _font(10)
                row += 1

        elif block.type == BlockType.TABLE:
            row = _write_table_block(ws, block, brand, row)
            row += 1

        elif block.type == BlockType.KPI:
            ws.cell(row, 1, block.kpi_value).font = Font(
                size=14, bold=True, color=brand.primary.lstrip('#')
            )
            ws.cell(row, 1).fill = _fill("#FFF0E0")
            if block.kpi_label:
                ws.cell(row, 2, block.kpi_label).font = _font(10, color=brand.text_light)
            row += 1

        else:
            row += 1

    # 열 너비 자동 조정
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 22


def _write_table_block(ws: Worksheet, block: Block, brand: BrandTheme, start_row: int) -> int:
    rows = [r for r in block.rows if r.cells]
    if not rows:
        return start_row

    cols = max(len(r.cells) for r in rows)
    row = start_row

    for r_idx, table_row in enumerate(rows):
        for c_idx, cell_text in enumerate(table_row.cells[:cols]):
            cell = ws.cell(row, c_idx + 1, str(cell_text))
            cell.border = _thin_border()
            if table_row.is_header:
                cell.fill = _fill(brand.secondary)
                cell.font = Font(size=10, bold=True, color=brand.background.lstrip('#'))
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                if r_idx % 2 == 0:
                    cell.fill = _fill("#F9F9F9")
                cell.font = _font(10)
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        row += 1

    # 필터 + 고정 (헤더가 있을 때)
    if rows and rows[0].is_header:
        ws.auto_filter.ref = (
            f"A{start_row}:{get_column_letter(cols)}{row - 1}"
        )
        ws.freeze_panes = ws.cell(start_row + 1, 1)

    return row


def generate(doc: DocumentContent, brand: BrandTheme,
             template: str = "internal",
             output_path: str = None) -> str:
    """
    DocumentContent + BrandTheme → XLSX 파일 생성
    """
    wb = Workbook()

    # 커버 시트
    _write_cover_sheet(wb, doc, brand)

    # 섹션별 시트
    for idx, sec in enumerate(doc.sections, 1):
        if not sec.blocks:
            continue
        _write_section_sheet(wb, sec, brand, idx)

    # 저장
    if output_path is None:
        base = os.path.splitext(os.path.basename(doc.source_path))[0]
        output_path = os.path.join(
            os.path.dirname(doc.source_path), "exports",
            f"{base}_{template}.xlsx"
        )

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path
